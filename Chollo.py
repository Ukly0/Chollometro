import time
import random
import requests
import openpyxl
import time
import random
import os
import requests
import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup

# Ruta del controlador del navegador (ejemplo para Chrome)
webdriver_service = Service('/path/to/chromedriver')

delay_min = 5
delay_max = 15

datanueva = []
def scrape():
    retardo = random.uniform(delay_min, delay_max)

    # Configuración del controlador del navegador
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')  # Ejecutar en modo sin cabeza (sin interfaz gráfica)
    options.add_argument('--no-sandbox')

    # URL de la página web
    urls = ['https://www.chollometro.com/populares','https://www.chollometro.com/']

    for url in urls:
        driver = webdriver.Chrome(service=webdriver_service, options=options)
    # Cargar la página
        driver.get(url)

        # Definir la función de desplazamiento mediante JavaScript
        def scroll_to_bottom():
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        # Realizar el desplazamiento hasta que no haya más contenido cargado
        while True:
            # Obtener el contenido actualizado después del desplazamiento
            updated_content = driver.page_source

            # Desplazarse hasta el final de la página
            scroll_to_bottom()

            # Esperar a que se cargue el nuevo contenido
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//footer")))

            # Verificar si se ha llegado al final de la página
            if updated_content == driver.page_source:
                break

        # Cerrar el controlador del navegador
        driver.quit()

        # Crear el objeto BeautifulSoup con el contenido actualizado
        soup = BeautifulSoup(updated_content, 'html.parser')

        # Buscar todos los elementos que contengan la clase deseada
        elements = soup.find_all(class_='thread cept-thread-item thread--type-list imgFrame-container--scale thread--deal')
        conprecio = True
        # Recorrer e imprimir los elementos encontrados
        for element in elements :
            
            tienda = element.find(class_='cept-merchant-name text--b text--color-brandPrimary link')
            if tienda:
                if tienda.text == "Amazon":
                    conprecio = True
                    time.sleep(retardo)

                    print(f'******** Nuevo producto de Amazon encontrado {tienda.text} **********')
                    titulo = element.find(class_= 'thread-title').find('a').text
                    linkid = element.find(class_= 'cept-tt thread-link linkPlain thread-title--list js-thread-title')['href']
                    
                    print(titulo)
                    id_Producto = ""
                    if element.find(class_= 'cept-vote-temp vote-temp vote-temp--hot'):
                        temperatura = element.find(class_= 'flex boxAlign-ai--all-c boxAlign-jc--all-sb space--b-3 space--fromW3-b-2').find('span').text
                        temperatura = int(temperatura.replace("°", ""))

                        if element.find(class_ = 'thread-price text--b cept-tp size--all-l size--fromW3-xl') is not None and temperatura >= 200:
                            precio = element.find(class_ = 'thread-price text--b cept-tp size--all-l size--fromW3-xl').text 
                        else:
                            conprecio = False
                            
                        if conprecio == True:   
                            img = element.find(class_='thread-image width--all-auto height--all-auto imgFrame-img')['src']
                            id= linkid.rsplit("-", 1)[-1]

                            if element.find(class_='mute--text text--lineThrough size--all-l size--fromW3-xl') is not None:
                                precio_original = element.find(class_ = 'mute--text text--lineThrough size--all-l size--fromW3-xl').text
                            else:
                                precio_original = "0"
                            
                            #Tratamiento del enlace
                            link = requests.get('https://www.chollometro.com/visit/threadmain/'+id)
                            linkAmazon = link.url

                            id_Producto_match = re.search(r'(?:dp/([\w]*))|(?:gp/product/([\w]*))', linkAmazon.split(" ")[0])

                            if id_Producto_match and tienda.text== "Amazon":
                                if id_Producto_match.group(1):
                                    id_Producto = id_Producto_match.group(1)
                                else:
                                    id_Producto = id_Producto_match.group(2)
                                print(id_Producto)
        
                                linkAfiliado = f"https://www.amazon.es/dp/{id_Producto}?th=1&psc=1&linkCode=ll1&tag=ukiy0-21&ref_=as_li_ss_tl"
                                print(linkAfiliado)
                            else:
                                print("---- ID DE PRODUCTO NO ENCONTRADO EN AMAZON")
                                                
                            print("+ Titulo: "+titulo)
                            print("+ Imagen: "+img)
                            print("+ Id: "+id)
                            print("+ Precio Anterior: "+precio_original)
                            print("+ Precio: "+precio)
                            print("+ Id_Producto: "+id_Producto)
                            print("+ Url: "+linkAfiliado)
                            print("+ Tienda: "+tienda.text)

                            ruta_archivo = os.path.join("Imagenes", id+".jpg")
                            response = requests.get(img)

                            with open(ruta_archivo, "wb") as archivo:
                                archivo.write(response.content)

                            datanueva.append([titulo, linkAfiliado, precio_original,precio, img, id, id_Producto])
                    else:
                        print ("---- TEMPERATURA NO ENONCTRADA")

    historico_archivo = 'historico_chollos.xlsx'
    nuevo_archivo = 'nuevos_chollos.xlsx'

    # Cargar el libro de Excel
    historico_excel = openpyxl.load_workbook(historico_archivo)

    # Obtener la hoja de trabajo activa
    historico_hoja_trabajo = historico_excel.active
    ultima_fila = historico_hoja_trabajo.max_row
    print('Ultima fila: '+str(ultima_fila))
    # Crear un array para almacenar los datos
    datosh = []

    # Recorrer las filas de la hoja de trabajo y leer los datos
    for fila in historico_hoja_trabajo.iter_rows(min_row=2, values_only=True):
        titulo, link, precio_original,precio, img, id,id_Producto= fila
        datosh.append([titulo, link, precio_original, precio, img, id,id_Producto])

    # Crear un nuevo libro de Excel y seleccionar la hoja de trabajo activa
    nuevo_archivo  = openpyxl.Workbook()
    nueva_hoja_trabajo = nuevo_archivo.active

    # Escribir los encabezados en la primera fila de la hoja de trabajo
    nueva_hoja_trabajo.append(["Titulo", "Link","Precio_Anterio", "Precio", "Imagen", "ID", "ID_Producto"])
    esta = False
    # Escribir los datos en la hoja de trabajo
    for fila in datanueva:
        esta = False
        for d in datosh:
            if d[5] == fila[5] or d[6] == fila[6]:
                esta = True
        if not esta:
            nueva_hoja_trabajo.append(fila)  
            ultima_fila += 1 
            historico_hoja_trabajo.append(fila)

    # Guardar el libro de Excel
    nombre_archivo = 'nuevos_chollos.xlsx'
    nuevo_archivo.save(nombre_archivo)
    historico_excel.save(historico_archivo)




